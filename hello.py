#ライブラリをインポート
from bs4 import BeautifulSoup
import requests
import openpyxl

#調べたいデータを指定
code = 3479
year = 2019

#URLを取得
url = "https://kabuoji3.com/stock/"+str(3479)+"/"+str(2019)+"/"
headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.116 Safari/537.36"}
soup = BeautifulSoup(requests.get(url, headers = headers).content,'html.parser')
title = soup.select_one("span.jp").text

#Excelファイルを作成
wb = openpyxl.Workbook()
ws = wb.active
ws.title = str(title)
ws['A1'].value = '日付'
ws['B1'].value = '始値'
ws['C1'].value = '高値'
ws['D1'].value = '安値'
ws['E1'].value = '終値'
ws['F1'].value = '出来高'
ws['G1'].value = '終値調整'

#株価を取り出す
all_tr = soup.find_all('tr')
for i in range(1,len(all_tr)):
    tr = all_tr[i].find_all('td') #リスト型
    for n,td in enumerate(tr, 1):
        new_cell = ws.cell(row=(i+1), column=n)
        new_cell.value = td.text

#ワークブックをExcelファイルとして保存
wb.save(str(title)+'.csv')